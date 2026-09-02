import os
import re
import tempfile
import email
from email import policy
import urllib.parse
import streamlit as st
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

st.set_page_config(
    page_title="AM PRO Racing System",
    page_icon="🏇",
    layout="wide"
)

st.markdown("""
    <style>
    .main-title {
        text-align: center;
        font-weight: 800;
        color: #1E3A8A;
        font-size: 26px;
        margin-bottom: 2px;
    }
    .sub-title {
        text-align: center;
        color: #4B5563;
        font-size: 14px;
        margin-bottom: 20px;
    }
    .stButton>button {
        width: 100%;
        border-radius: 8px;
        height: 3.2em;
        background-color: #2563EB;
        color: white;
        font-weight: bold;
        font-size: 16px;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown("<div class='main-title'>🏇 AM PRO Racing System</div>", unsafe_allow_html=True)
st.markdown("<div class='sub-title'>Step 1: All Form Matrix | Step 2: AM Final Score & Logic Engine</div>", unsafe_allow_html=True)

# =====================================================================
# HELPER & CLEANING FUNCTIONS
# =====================================================================

def extract_html_from_bytes(file_bytes):
    try:
        msg = email.message_from_bytes(file_bytes, policy=policy.default)
        if msg.is_multipart():
            for part in msg.iter_parts():
                if part.get_content_type() == "text/html":
                    return part.get_content()
    except Exception:
        pass

    text = file_bytes.decode("utf-8", errors="ignore")
    match = re.search(r"(<!DOCTYPE\s+html[^>]*>[\s\S]*?</html>)", text, re.I)
    if match:
        return match.group(1)
    return text

def clean_jockey_name(name):
    if not name:
        return ""
    cleaned = re.sub(r"\(a\d*(\.\d+)?(/[0-9.]+kg)?\)", "", str(name), flags=re.I)
    cleaned = re.sub(r"\(a\)", "", cleaned, flags=re.I)
    cleaned = re.sub(r"\(cd\s*[\d.]+kg\)", "", cleaned, flags=re.I)
    cleaned = re.sub(r"^(Ms|Mr|Mrs)\s+", "", cleaned.strip(), flags=re.I)
    cleaned = re.sub(r"\s+", " ", cleaned).strip().lower()
    return cleaned

def extract_weight_num(w_str):
    if not w_str:
        return 0.0
    match = re.search(r"(\d+(\.\d+)?)", str(w_str))
    return float(match.group(1)) if match else 0.0

def normalize_track(t_str):
    if not t_str:
        return ""
    t_clean = str(t_str).strip().lower()
    if "good" in t_clean: return "good"
    if "soft" in t_clean: return "soft"
    if "heavy" in t_clean: return "heavy"
    if "synthetic" in t_clean or "syn" in t_clean or "tapeta" in t_clean: return "synthetic"
    if "firm" in t_clean: return "firm"
    return t_clean

def normalize_code(c_str):
    if not c_str:
        return ""
    return re.sub(r"[\s\-_.]", "", str(c_str)).strip().lower()

def parse_time_to_seconds(t_str):
    if not t_str or not isinstance(t_str, str):
        return 0.0
    t_str = t_str.strip()
    try:
        if ":" in t_str:
            p = t_str.split(":")
            return (float(p[0]) * 60.0) + float(p[1])
        return float(t_str)
    except Exception:
        return 0.0

def format_seconds_to_time(secs):
    if not secs or secs <= 0:
        return ""
    m = int(secs // 60)
    s = secs % 60
    if m > 0:
        return f"{m}:{s:05.2f}"
    return f"0:{s:05.2f}"

def clean_filename(name):
    clean = re.sub(r'[\\/*?:"<>|]', "", str(name))
    clean = clean.replace("%20", " ").strip()
    return clean

# =====================================================================
# HTML ALL FORM PARSER
# =====================================================================

def parse_racing_australia_html(html_text):
    soup = BeautifulSoup(html_text, "html.parser")
    races = []

    meeting_venue = "Warwick Farm"
    meeting_country = "NSW"
    date_formatted = "02Sep2026"

    venue_elem = soup.find("div", class_="race-venue")
    if venue_elem:
        h2 = venue_elem.find("h2")
        if h2:
            meeting_venue = h2.get_text().split(":")[0].strip()
        date_span = venue_elem.find("span", class_="race-venue-date")
        if date_span:
            meeting_date_raw = date_span.get_text().strip()
            d_match = re.search(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", meeting_date_raw)
            if d_match:
                day, month, year = d_match.groups()
                date_formatted = f"{int(day):02d}{month[:3]}{year}"

    banner = soup.find("div", class_="state-specific-banner")
    if banner and banner.find("img"):
        alt_txt = banner.find("img").get("alt", "")
        m_state = re.search(r"\b(NSW|VIC|QLD|WA|SA|TAS|ACT|NT)\b", alt_txt, re.I)
        if m_state:
            meeting_country = m_state.group(1).upper()

    race_titles = soup.find_all("table", class_="race-title")

    for idx, r_title_table in enumerate(race_titles, start=1):
        race_num_str = f"Race {idx}"
        sheet_name_str = f"R{idx}"

        title_text = r_title_table.get_text()
        dist_match = re.search(r"\((\d{3,4})\s*METRES\)", title_text, re.I)
        race_dist = f"{dist_match.group(1)}m" if dist_match else "1200m"

        class_match = re.search(r"\b(BenchMark\s*\d+|BM\s*\d+|Maiden|MDN|Open|Handicap|Listed|G[1-3])\b", title_text, re.I)
        race_class = class_match.group(1) if class_match else "Handicap"

        fields_table = r_title_table.find_next("table", class_="race-strip-fields")
        current_horses_map = {}

        if fields_table:
            for tr in fields_table.find_all("tr"):
                if "Scratched" in tr.get("class", []):
                    continue
                no_td = tr.find("td", class_="no")
                horse_td = tr.find("td", class_="horse")
                jockey_td = tr.find("td", class_="jockey")
                trainer_td = tr.find("td", class_="trainer")
                barrier_td = tr.find("td", class_="barrier")
                weight_td = tr.find("td", class_="weight")

                if no_td and horse_td:
                    h_no = no_td.get_text(strip=True)
                    h_name = re.sub(r"\s*\(.*?\)", "", horse_td.get_text(strip=True)).strip().upper()
                    jock_name = jockey_td.get_text(strip=True) if jockey_td else ""
                    train_name = trainer_td.get_text(strip=True) if trainer_td else ""
                    barrier_val = barrier_td.get_text(strip=True) if barrier_td else ""
                    weight_val = weight_td.get_text(strip=True) if weight_td else "58.0kg"

                    current_horses_map[h_name] = {
                        "no": h_no,
                        "name": h_name,
                        "jockey": jock_name,
                        "trainer": train_name,
                        "owner": "",
                        "barrier": barrier_val,
                        "final_weight": weight_val,
                        "stats": "",
                        "runs": []
                    }

        race_horses_list = []
        all_horse_forms = soup.find_all("table", class_="horse-form-table")

        for h_form in all_horse_forms:
            if h_form.find("td", class_="Scratched"):
                continue

            h_name_tag = h_form.find("span", class_="horse-name")
            h_no_tag = h_form.find("span", class_="horse-number")

            if not h_name_tag:
                continue

            raw_h_name = h_name_tag.get_text(strip=True).upper()
            h_clean_name = re.sub(r"\s*\(.*?\)", "", raw_h_name).strip().upper()
            h_num = h_no_tag.get_text(strip=True) if h_no_tag else ""

            if h_clean_name not in current_horses_map:
                continue

            h_obj = current_horses_map[h_clean_name]

            form_text = h_form.get_text(separator=" ", strip=True)
            stats_dict = {}
            for stat_key in ["1st Up", "2nd Up", "Track", "Dist", "Track/Dist", "Firm", "Good", "Soft", "Heavy", "Synthetic"]:
                sm = re.search(rf"\b{re.escape(stat_key)}:\s*([\d:-]+)", form_text)
                stats_dict[stat_key] = sm.group(1) if sm else "0:0-0-0"

            h_obj["stats"] = ",".join([
                stats_dict.get("1st Up", "0:0-0-0"),
                stats_dict.get("2nd Up", "0:0-0-0"),
                stats_dict.get("Track", "0:0-0-0"),
                stats_dict.get("Dist", "0:0-0-0"),
                stats_dict.get("Track/Dist", "0:0-0-0"),
                stats_dict.get("Firm", "0:0-0-0"),
                stats_dict.get("Good", "0:0-0-0"),
                stats_dict.get("Soft", "0:0-0-0"),
                stats_dict.get("Heavy", "0:0-0-0"),
                stats_dict.get("Synthetic", "0:0-0-0")
            ])

            owner_m = re.search(r"Owners:\s*(.*?)(?:Colours:|Gear Changes:|<br|$)", form_text)
            if owner_m and not h_obj["owner"]:
                h_obj["owner"] = owner_m.group(1).strip()

            last_start_tbl = h_form.find("table", class_="horse-last-start")
            if last_start_tbl:
                for r_tr in last_start_tbl.find_all("tr"):
                    pos_td = r_tr.find("td", class_="Pos")
                    remain_td = r_tr.find("td", class_="remain")

                    if not remain_td:
                        continue

                    pos_val = pos_td.get_text(separator=" ", strip=True) if pos_td else ""
                    rem_text = remain_td.get_text(separator=" ", strip=True)

                    run_type = "Race"
                    if "Jump Out" in rem_text or pos_val.startswith("J"):
                        run_type = "Jump Out"
                    elif "-BT" in rem_text or "-TRL" in rem_text or pos_val.startswith("T"):
                        run_type = "Trial"

                    venue_date_m = re.search(r"^([A-Z\s.]+?)\s+(\d{1,2}[A-Za-z]{3}\d{2,4})", rem_text)
                    if venue_date_m:
                        p_place = venue_date_m.group(1).strip()
                        p_date = venue_date_m.group(2).strip()
                    else:
                        p_place = ""
                        p_date = ""

                    d_m = re.search(r"\b(\d{3,4}m)\b", rem_text)
                    p_dist = d_m.group(1) if d_m else ""

                    tr_m = re.search(r"\b(Good\s*\d?|Soft\s*\d?|Heavy\s*\d?|Synthetic|Firm\s*\d?|Fast)\b", rem_text, re.I)
                    p_track = tr_m.group(1) if tr_m else "Good4"

                    cl_m = re.search(r"\b(BM\d+|MDN-SW|2Y-BT|OPEN-BT|2Y\s+MDN|3Y\s+MDN|SUPER\s+3Y\s+MDN|2YF-BT|2YC&G-BT|3Y\+MDN-BT|CL\d+|LR|G[1-3])\b", rem_text, re.I)
                    p_class = cl_m.group(1) if cl_m else ("Trial" if run_type == "Trial" else "")

                    w_m = re.search(r"\b(\d{2}(?:\.\d)?kg)\b", rem_text, re.I)
                    p_weight = w_m.group(1) if w_m else ("0kg" if run_type in ["Trial", "Jump Out"] else "56kg")

                    bar_m = re.search(r"Barrier\s*(\d{1,2})", rem_text, re.I)
                    p_barrier = bar_m.group(1) if bar_m else "0"

                    jock_tag = remain_td.find("a", href=re.compile(r"JockeyLastRuns", re.I))
                    p_jockey = jock_tag.get_text(strip=True) if jock_tag else ""

                    time_m = re.search(r"\b(\d{1,2}:\d{2}\.\d{2})\b", rem_text)
                    p_time = time_m.group(1) if time_m else ""

                    p800_m = re.search(r"(\d{1,2}(?:st|nd|rd|th)@800m)", rem_text)
                    p400_m = re.search(r"(\d{1,2}(?:st|nd|rd|th)@400m)", rem_text)
                    p_800 = p800_m.group(1) if p800_m else ""
                    p_400 = p400_m.group(1) if p400_m else ""

                    split_m = re.search(r"\(600m\s*([\d.]+)\)", rem_text)
                    p_split = split_m.group(1) if split_m else ""

                    odds_m = re.search(r"(\$[\d/.]+[Ff]?)$", rem_text)
                    p_odds = odds_m.group(1) if odds_m else ("$000" if run_type in ["Trial", "Jump Out"] else "")

                    calc_time_val = ""
                    if p_time and p_dist:
                        try:
                            past_dist_num = float(re.sub(r"\D", "", p_dist))
                            curr_dist_num = float(re.sub(r"\D", "", race_dist))
                            base_secs = parse_time_to_seconds(p_time)

                            if past_dist_num > 0 and base_secs > 0:
                                scaled_secs = base_secs * (curr_dist_num / past_dist_num)
                                cur_w_num = extract_weight_num(h_obj["final_weight"])
                                past_w_num = extract_weight_num(p_weight)
                                if cur_w_num and past_w_num:
                                    w_diff = cur_w_num - past_w_num
                                    scaled_secs += (w_diff * 0.15)
                                calc_time_val = format_seconds_to_time(scaled_secs)
                        except Exception:
                            calc_time_val = ""

                    h_obj["runs"].append({
                        "date": p_date,
                        "place": p_place,
                        "type": run_type,
                        "pos": pos_val,
                        "class": p_class,
                        "distance": p_dist,
                        "track": p_track,
                        "barrier": p_barrier,
                        "weight": p_weight,
                        "jockey": p_jockey,
                        "trainer": h_obj["trainer"],
                        "owner": h_obj["owner"],
                        "rating": "",
                        "p800": p_800,
                        "p400": p_400,
                        "split": p_split,
                        "time": p_time,
                        "odds": p_odds,
                        "calc_time": calc_time_val
                    })

            race_horses_list.append(h_obj)

        if race_horses_list:
            races.append({
                "race_no": race_num_str,
                "sheet_name": sheet_name_str,
                "country": meeting_country,
                "distance": race_dist,
                "track_condition": "Good 4",
                "class_name": race_class,
                "place": meeting_venue,
                "date_formatted": date_formatted,
                "horses": race_horses_list
            })

    return races

# =====================================================================
# EXACT MULTI-SHEET EXCEL BUILDER (INCL. AM SCORE HEADERS L, M, N, O, P, Q)
# =====================================================================

def generate_am_pro_step1_workbook(races):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    dark_gray_fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    white_bold = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    regular_font = Font(name="Arial", size=9)
    bold_font = Font(name="Arial", size=9, bold=True)

    yellow_match_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bold_green_font = Font(name="Arial", size=9, bold=True, color="006100")
    purple_font = Font(name="Arial", size=9, bold=True, italic=True, color="7E22CE")

    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    first_sheet_title = races[0]["sheet_name"] if races else "R1"

    for r_idx, race in enumerate(races, start=1):
        ws = wb.create_sheet(title=race["sheet_name"])
        cur_dist = race["distance"]
        cur_track = race["track_condition"]

        # 1. Top Metadata
        ws["A3"] = "Country"
        ws["B3"] = race["country"]
        ws["A4"] = "Place"

        if r_idx == 1:
            ws["B4"] = race["place"]
        else:
            ws["B4"] = f"='{first_sheet_title}'!B4"

        ws["A5"] = "Distance"
        ws["B5"] = cur_dist
        ws["A6"] = "Track Condition"
        ws["B6"] = cur_track
        ws["A7"] = "Class"
        ws["B7"] = race["class_name"]
        ws["A8"] = "Horse Count"
        ws["B8"] = len(race["horses"])

        for row_i in range(3, 9):
            ws.cell(row=row_i, column=1).font = bold_font
            ws.cell(row=row_i, column=2).font = regular_font

        # 2. Current Race Table (Row 11) - INCL AM SCORE HEADERS L to Q
        current_headers = [
            "Horse NO", "Horse Name", "Jockey Name", "Trainer Name", "Owner Name",
            "Barrier", "Final Weight", "Distance", "Track Condition", "Class", "Rating",
            "AM Score", "Cond 1: Dist/Track/Class", "Cond 2: Jockey 70%", "Cond 3: Class Drop & Jockey",
            "Cond 4: Weight Drop", "Cond 5 & 6: Venue/Season Form"
        ]

        ws.row_dimensions[11].height = 28
        for col_idx, h_text in enumerate(current_headers, start=1):
            c = ws.cell(row=11, column=col_idx, value=h_text)
            c.fill = navy_fill
            c.font = white_bold
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        cur_row = 12
        horse_jockey_map = {}
        horse_weight_map = {}
        horse_barrier_map = {}

        for h in race["horses"]:
            horse_jockey_map[h["name"].lower()] = clean_jockey_name(h["jockey"])
            horse_weight_map[h["name"].lower()] = extract_weight_num(h["final_weight"])
            horse_barrier_map[h["name"].lower()] = str(h["barrier"]).strip()

            row_data = [
                h["no"], h["name"], h["jockey"], h["trainer"], h["owner"],
                h["barrier"], h["final_weight"], cur_dist, cur_track, race["class_name"], "",
                "", "", "", "", "", ""
            ]
            ws.append(row_data)

            for col_idx in range(1, len(current_headers) + 1):
                c = ws.cell(row=cur_row, column=col_idx)
                c.font = regular_font
                c.border = border_thin
                c.alignment = Alignment(vertical="center")
            cur_row += 1

        # 3. Previous Runs Section
        r_ptr = cur_row + 3

        prev_headers = [
            "Date", "Place", "Run Type", "Finishing Position", "Class", "Distance", "Track Condition",
            "Barrier", "Weight", "Weight Diff", "Jockey Name", "Trainer Name", "Owner Name",
            "Rating", "Position @800", "Position @400", "600m Split", "Finishing Time", "Odds", "Calculated Time"
        ]

        first_past_run_row = r_ptr + 1

        for h in race["horses"]:
            ws.cell(row=r_ptr, column=1, value=h["name"]).font = bold_font
            ws.cell(row=r_ptr, column=3, value=h["no"]).font = bold_font
            ws.cell(row=r_ptr, column=4, value=h["jockey"]).font = bold_font
            r_ptr += 1

            stats_headers = ["Min/Max-Dist-Win", "1st Up", "2nd Up", "Track", "Dist", "Track/Dist", "Firm", "Good", "Soft", "Heavy", "Synthetic"]
            for s_i, s_h in enumerate(stats_headers, start=1):
                ws.cell(row=r_ptr, column=s_i, value=s_h).font = bold_font
            r_ptr += 1

            ws.cell(row=r_ptr, column=1, value="")
            stat_parts = h.get("stats", "").split(",")
            for s_i, s_v in enumerate(stat_parts, start=2):
                if s_i <= 11:
                    ws.cell(row=r_ptr, column=s_i, value=s_v).font = regular_font
            r_ptr += 2

            ws.row_dimensions[r_ptr].height = 22
            for col_idx, h_text in enumerate(prev_headers, start=1):
                c = ws.cell(row=r_ptr, column=col_idx, value=h_text)
                c.fill = dark_gray_fill
                c.font = white_bold
                c.alignment = Alignment(horizontal="center", vertical="center")
            r_ptr += 1

            runs = h.get("runs", [])
            for run in runs:
                cur_wgt_val = horse_weight_map.get(h["name"].lower(), 0.0)
                past_wgt_val = extract_weight_num(run.get("weight", "0"))
                wgt_diff = cur_wgt_val - past_wgt_val if (cur_wgt_val and past_wgt_val) else 0.0
                wgt_diff_str = f"{wgt_diff:+.1f}kg" if wgt_diff != 0 else "0.0kg"

                row_vals = [
                    run.get("date", ""), run.get("place", ""), run.get("type", "Race"),
                    run.get("pos", ""), run.get("class", ""), run.get("distance", ""), run.get("track", ""),
                    run.get("barrier", ""), run.get("weight", ""), wgt_diff_str,
                    run.get("jockey", ""), h["trainer"], h["owner"],
                    "", run.get("p800", ""), run.get("p400", ""), run.get("split", ""),
                    run.get("time", ""), run.get("odds", "$000"),
                    run.get("calc_time", "")
                ]
                ws.append(row_vals)

                for c_i in range(1, len(prev_headers) + 1):
                    cell = ws.cell(row=r_ptr, column=c_i)
                    cell.border = border_thin
                    cell.font = regular_font

                pos_str = str(run.get("pos", "")).lower()
                if "1 of" in pos_str or "2 of" in pos_str or "3 of" in pos_str or pos_str in ["1", "2", "3"]:
                    ws.cell(row=r_ptr, column=4).fill = green_win_fill
                    ws.cell(row=r_ptr, column=4).font = bold_green_font

                if re.sub(r"\D", "", run.get("distance", "")) == re.sub(r"\D", "", cur_dist):
                    ws.cell(row=r_ptr, column=6).fill = yellow_match_fill

                if normalize_track(run.get("track", "")) == normalize_track(cur_track):
                    ws.cell(row=r_ptr, column=7).fill = yellow_match_fill

                r_ptr += 1

            r_ptr += 3
            ws.append([])
            ws.append([])
            ws.append([])

        last_past_run_row = max(r_ptr, 120)

        cf_range = f"K{first_past_run_row}:K{last_past_run_row}"
        purple_font_rule = FormulaRule(
            formula=[f'AND(K{first_past_run_row}<>"", COUNTIF($R$1:$R$24, "*"&TRIM(K{first_past_run_row})&"*")>0)'],
            font=purple_font
        )
        ws.conditional_formatting.add(cf_range, purple_font_rule)

        for col_letter in ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T"]:
            ws.column_dimensions[col_letter].width = 8.5

    return wb

# =====================================================================
# STEP 2: PROCESS EDITED EXCEL (AM SCORE CALCULATION & 6 CONDITIONS LOGIC)
# =====================================================================

def process_step2_edited_excel(wb):
    yellow_match_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bold_green_font = Font(name="Arial", size=9, bold=True, color="006100")
    purple_font = Font(name="Arial", size=9, bold=True, italic=True, color="7E22CE")

    first_ws = wb.worksheets[0]
    r1_master_place = str(first_ws["B4"].value or "").strip()
    r1_master_place = re.sub(r"^[='\"+]+", "", r1_master_place).strip()
    r1_master_norm = normalize_code(r1_master_place)

    for ws in wb.worksheets:
        cur_dist = str(ws["B5"].value or "").strip()
        cur_track = str(ws["B6"].value or "").strip()
        cur_class_b7 = str(ws["B7"].value or "").strip()
        cur_class_norm = normalize_code(cur_class_b7)
        cur_dist_num = extract_weight_num(cur_dist)

        target_jockeys_r = set()
        for r_idx in range(1, 25):
            val = str(ws.cell(row=r_idx, column=18).value or "").strip()
            if val and val.lower() != "finishing time":
                cleaned_j = clean_jockey_name(val)
                if cleaned_j:
                    target_jockeys_r.add(cleaned_j)

        # Build Horse Runs Map for AM Score Evaluation
        horses_eval_data = {}
        row_idx_map = {}

        for r in range(12, ws.max_row + 1):
            h_no = str(ws.cell(row=r, column=1).value or "").strip()
            h_name = str(ws.cell(row=r, column=2).value or "").strip().upper()
            jock = str(ws.cell(row=r, column=3).value or "").strip()
            barrier = str(ws.cell(row=r, column=6).value or "").strip()
            wgt = str(ws.cell(row=r, column=7).value or "").strip()

            if not h_no or not h_no.isdigit():
                break

            row_idx_map[h_name] = r
            horses_eval_data[h_name] = {
                "jockey": clean_jockey_name(jock),
                "barrier": barrier,
                "weight": extract_weight_num(wgt),
                "runs": []
            }

        current_eval_horse = None
        for r in range(15, ws.max_row + 1):
            c1_val = str(ws.cell(row=r, column=1).value or "").strip().upper()
            c4_val = str(ws.cell(row=r, column=4).value or "").strip()

            if c1_val in horses_eval_data:
                current_eval_horse = c1_val
                if c4_val:
                    horses_eval_data[c1_val]["jockey"] = clean_jockey_name(c4_val)
                continue

            date_cell = ws.cell(row=r, column=1)
            place_cell = ws.cell(row=r, column=2)
            pos_cell = ws.cell(row=r, column=4)
            class_cell = ws.cell(row=r, column=5)
            dist_cell = ws.cell(row=r, column=6)
            track_cell = ws.cell(row=r, column=7)
            barrier_cell = ws.cell(row=r, column=8)
            weight_cell = ws.cell(row=r, column=9)
            jockey_cell = ws.cell(row=r, column=11)

            pos_str = str(pos_cell.value or "").lower()

            if not date_cell.value and not pos_cell.value:
                continue
            if str(date_cell.value or "").lower() == "date":
                continue

            # Place match B4
            if place_cell.value and r1_master_norm:
                past_place_norm = normalize_code(place_cell.value)
                if r1_master_norm == past_place_norm or r1_master_norm in past_place_norm or past_place_norm in r1_master_norm:
                    place_cell.fill = yellow_match_fill

            # Win/Place Top 3
            is_top3 = any(p in pos_str for p in ["1 of", "2 of", "3 of", "1", "2", "3", "1st", "2nd", "3rd"])
            is_top2 = any(p in pos_str for p in ["1 of", "2 of", "1", "2", "1st", "2nd"])
            is_top4 = is_top3 or "4 of" in pos_str or "4" in pos_str or "4th" in pos_str

            if is_top3:
                pos_cell.fill = green_win_fill
                pos_cell.font = bold_green_font

            # Class match B7
            if class_cell.value and cur_class_norm:
                past_class_norm = normalize_code(class_cell.value)
                if cur_class_norm == past_class_norm or cur_class_norm in past_class_norm or past_class_norm in cur_class_norm:
                    class_cell.fill = yellow_match_fill

            # Distance match
            if dist_cell.value and cur_dist:
                if re.sub(r"\D", "", str(dist_cell.value)) == re.sub(r"\D", "", cur_dist):
                    dist_cell.fill = yellow_match_fill

            # Track match
            if track_cell.value and cur_track:
                if normalize_track(str(track_cell.value)) == normalize_track(cur_track):
                    track_cell.fill = yellow_match_fill

            # Barrier match
            if current_eval_horse and barrier_cell.value:
                cur_bar = horses_eval_data[current_eval_horse]["barrier"]
                past_bar = str(barrier_cell.value).strip()
                if cur_bar and past_bar and (cur_bar == past_bar) and past_bar != "0":
                    barrier_cell.fill = yellow_match_fill

            # Jockey highlight
            if jockey_cell.value:
                past_jock_clean = clean_jockey_name(str(jockey_cell.value))
                matched_r = False
                for tj in target_jockeys_r:
                    if tj in past_jock_clean or past_jock_clean in tj:
                        jockey_cell.font = purple_font
                        matched_r = True
                        break

                if not matched_r and current_eval_horse:
                    cur_jock = horses_eval_data[current_eval_horse]["jockey"]
                    if cur_jock and past_jock_clean and (cur_jock in past_jock_clean or past_jock_clean in cur_jock):
                        jockey_cell.fill = yellow_match_fill

            # Collect run for AM score
            if current_eval_horse:
                horses_eval_data[current_eval_horse]["runs"].append({
                    "dist": str(dist_cell.value or ""),
                    "track": str(track_cell.value or ""),
                    "class": str(class_cell.value or ""),
                    "pos": pos_str,
                    "jockey": clean_jockey_name(str(jockey_cell.value or "")),
                    "weight": extract_weight_num(weight_cell.value),
                    "place": str(place_cell.value or "")
                })

        # =====================================================================
        # CALCULATE AM SCORES & WRITE TO COLUMNS L TO Q
        # =====================================================================
        for h_name, h_info in horses_eval_data.items():
            r_idx = row_idx_map.get(h_name)
            if not r_idx:
                continue

            cur_jock = h_info["jockey"]
            cur_wgt = h_info["weight"]
            runs = h_info["runs"]
            total_runs = len(runs)

            # Condition 1: Same Dist, Track, Class & Top 3
            cond1_score = 0
            cond1_desc = ""
            for run in runs:
                d_match = re.sub(r"\D", "", run["dist"]) == re.sub(r"\D", "", cur_dist)
                t_match = normalize_track(run["track"]) == normalize_track(cur_track)
                c_match = cur_class_norm and (cur_class_norm in normalize_code(run["class"]) or normalize_code(run["class"]) in cur_class_norm)
                is_top3 = any(p in run["pos"] for p in ["1 of", "2 of", "3 of", "1", "2", "3", "1st", "2nd", "3rd"])
                if d_match and t_match and c_match and is_top3:
                    cond1_score = 1
                    cond1_desc = "Cond 1: Same Dist/Track/Class Top 3 (+1)"
                    break

            # Condition 2: Same Jockey 70%+ Top 2
            cond2_score = 0
            cond2_desc = ""
            if total_runs > 0 and cur_jock:
                jock_runs = [r for r in runs if cur_jock in r["jockey"] or r["jockey"] in cur_jock]
                if len(jock_runs) >= 2:
                    jock_top2 = sum(1 for r in jock_runs if any(p in r["pos"] for p in ["1 of", "2 of", "1", "2", "1st", "2nd"]))
                    pct = (jock_top2 / len(jock_runs)) * 100.0
                    if pct >= 70.0:
                        cond2_score = 1
                        cond2_desc = f"Cond 2: Jockey {pct:.0f}% Top 2 (+1)"

            # Condition 3: Higher Class Experience (Top 4) -> Lower Class Top 2 (Same Track) + Same Jockey extra
            cond3_score = 0
            cond3_desc_parts = []
            has_higher_exp = any(any(hc in normalize_code(r["class"]) for hc in ["bm84", "bm90", "bm100", "open", "g1", "g2", "g3", "lr"]) and any(p in r["pos"] for p in ["1", "2", "3", "4"]) for r in runs)
            if has_higher_exp:
                for run in runs:
                    t_match = normalize_track(run["track"]) == normalize_track(cur_track)
                    is_top2 = any(p in run["pos"] for p in ["1 of", "2 of", "1", "2", "1st", "2nd"])
                    if t_match and is_top2:
                        cond3_score += 1
                        cond3_desc_parts.append("Class Drop Top 2 (+1)")
                        if cur_jock and (cur_jock in run["jockey"] or run["jockey"] in cur_jock):
                            cond3_score += 1
                            cond3_desc_parts.append("Same Jockey Bonus (+1)")
                        break
            cond3_desc = " | ".join(cond3_desc_parts)

            # Condition 4: Weight Drop >= 2.5kg
            cond4_score = 0
            cond4_desc = ""
            for run in runs:
                past_wgt = run["weight"]
                if past_wgt > 0 and cur_wgt > 0:
                    wgt_drop = past_wgt - cur_wgt
                    if wgt_drop >= 2.5:
                        cond4_score = 1
                        cond4_desc = f"Cond 4: Weight Drop {wgt_drop:.1f}kg (+1)"
                        break

            # Condition 5 & 6: Venue Place % (>=60% Top 3) & Recent 6M Form (>=60% Top 3)
            cond5_6_score = 0
            cond5_6_desc_parts = []
            if total_runs > 0 and r1_master_norm:
                venue_runs = [r for r in runs if r1_master_norm in normalize_code(r["place"]) or normalize_code(r["place"]) in r1_master_norm]
                if len(venue_runs) >= 2:
                    v_top3 = sum(1 for r in venue_runs if any(p in r["pos"] for p in ["1 of", "2 of", "3 of", "1", "2", "3", "1st", "2nd", "3rd"]))
                    if (v_top3 / len(venue_runs)) >= 0.60:
                        cond5_6_score += 1
                        cond5_6_desc_parts.append("Venue >60% Top3 (+1)")

                # Assuming recent form from runs order
                recent_runs = runs[:5] if len(runs) >= 5 else runs
                if len(recent_runs) >= 2:
                    r_top3 = sum(1 for r in recent_runs if any(p in r["pos"] for p in ["1 of", "2 of", "3 of", "1", "2", "3", "1st", "2nd", "3rd"]))
                    if (r_top3 / len(recent_runs)) >= 0.60:
                        cond5_6_score += 1
                        cond5_6_desc_parts.append("Recent Form >60% (+1)")
            cond5_6_desc = " | ".join(cond5_6_desc_parts)

            total_am_score = cond1_score + cond2_score + cond3_score + cond4_score + cond5_6_score

            # Write to Excel Columns L, M, N, O, P, Q
            ws.cell(row=r_idx, column=12, value=total_am_score if total_am_score > 0 else "") # L: AM Score
            ws.cell(row=r_idx, column=13, value=cond1_desc) # M
            ws.cell(row=r_idx, column=14, value=cond2_desc) # N
            ws.cell(row=r_idx, column=15, value=cond3_desc) # O
            ws.cell(row=r_idx, column=16, value=cond4_desc) # P
            ws.cell(row=r_idx, column=17, value=cond5_6_desc) # Q

        for col_letter in ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T"]:
            ws.column_dimensions[col_letter].width = 8.5

# =====================================================================
# STREAMLIT UI
# =====================================================================

tab1, tab2 = st.tabs(["🌐 STEP 1: Racing Australia Input (MHT / HTML / URL)", "📊 STEP 2: Process Edited Excel (AM Final Score)"])

with tab1:
    st.subheader("1. Racing Australia All Form-லிருந்து Excel உருவாக்குதல்")
    
    input_choice = st.radio(
        "உள்ளீட்டு முறையைத் தேர்ந்தெடுக்கவும் (Input Method):",
        [
            "📁 Upload Webpage File (.mht / .mhtml / .html)",
            "📋 Paste Page Source HTML (குறியீடு பேஸ்ட் செய்தல்)"
        ],
        horizontal=True
    )

    if input_choice == "📁 Upload Webpage File (.mht / .mhtml / .html)":
        st.caption("💡 **மொபைல் வழிமுறை:** Chrome-ல் Racing Australia AllForm பக்கத்தில் 3 புள்ளிகளைத் (︙) தட்டி **Download (⬇️)** கொடுத்தால் சேமிக்கப்படும் `.mht` அல்லது `.html` ஃபைலை இங்கே பதிவேற்றவும்.")
        uploaded_file = st.file_uploader(
            "Webpage ஃபைலை தேர்வு செய்யவும் (.mht, .mhtml, .html, .htm):",
            type=["mht", "mhtml", "html", "htm"],
            key="uploader_webpage_file"
        )
        
        if uploaded_file is not None:
            if st.button("🚀 Process File & Generate Excel", type="primary", key="btn_file_gen"):
                with st.spinner("🔄 ஃபைலிலிருந்து பந்தய விவரங்கள் பெறப்படுகின்றன..."):
                    raw_bytes = uploaded_file.read()
                    html_content = extract_html_from_bytes(raw_bytes)
                    races = parse_racing_australia_html(html_content)
                    
                    if not races:
                        st.error("❌ விவரங்களைப் பிரிக்க முடியவில்லை. சரியான AllForm பக்கம் தானா என சரிபார்க்கவும்.")
                    else:
                        total_horses = sum(len(r["horses"]) for r in races)
                        total_runs = sum(sum(len(h["runs"]) for h in r["horses"]) for r in races)

                        date_prefix = races[0].get("date_formatted", "02Sep2026")
                        b4_place = clean_filename(races[0]["place"])
                        out_filename = f"{date_prefix}_{b4_place}.xlsx"

                        wb = generate_am_pro_step1_workbook(races)
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                            temp_path = tmp.name
                        wb.save(temp_path)

                        with open(temp_path, "rb") as f:
                            excel_data = f.read()

                        try: os.remove(temp_path)
                        except Exception: pass

                        st.success(f"🎉 வெற்றி! {len(races)} பந்தயங்கள் | {total_horses} குதிரைகள் | {total_runs} முந்தைய ஓட்டங்கள் கண்டறியப்பட்டன.")
                        st.download_button(
                            label=f"📥 Download ({out_filename})",
                            data=excel_data,
                            file_name=out_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

    else:
        html_input = st.text_area(
            "📋 Racing Australia Page Source (HTML Code):",
            height=240,
            placeholder="<!DOCTYPE html>... (Ctrl + U கொடுத்து முழு குறியீட்டையும் பேஸ்ட் செய்யவும்)",
            key="exact_html_input_box"
        )
        if st.button("🚀 Process HTML & Generate Excel", type="primary", key="btn_exact_gen"):
            cleaned = html_input.strip()
            if not cleaned:
                st.warning("⚠️ தயவுசெய்து HTML குறியீட்டை பேஸ்ட் செய்யவும்.")
            else:
                with st.spinner("🔄 HTML-லிருந்து தரவுகள் பிரித்தெடுக்கப்படுகின்றன..."):
                    races = parse_racing_australia_html(cleaned)
                    if not races:
                        st.error("❌ HTML-லிருந்து தரவுகளைப் பிரிக்க முடியவில்லை.")
                    else:
                        total_horses = sum(len(r["horses"]) for r in races)
                        total_runs = sum(sum(len(h["runs"]) for h in r["horses"]) for r in races)

                        date_prefix = races[0].get("date_formatted", "02Sep2026")
                        b4_place = clean_filename(races[0]["place"])
                        out_filename = f"{date_prefix}_{b4_place}.xlsx"

                        wb = generate_am_pro_step1_workbook(races)
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                            temp_path = tmp.name
                        wb.save(temp_path)

                        with open(temp_path, "rb") as f:
                            excel_data = f.read()

                        try: os.remove(temp_path)
                        except Exception: pass

                        st.success(f"🎉 வெற்றி! {len(races)} பந்தயங்கள் | {total_horses} குதிரைகள் | {total_runs} முந்தைய ஓட்டங்கள் கண்டறியப்பட்டன.")
                        st.download_button(
                            label=f"📥 Download ({out_filename})",
                            data=excel_data,
                            file_name=out_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

with tab2:
    st.subheader("2. திருத்தப்பட்ட Multi-Sheet Excel-ஐப் பதிவேற்றி Final அறிக்கை பெறுதல்")
    st.caption("R1 ஷீட்டில் B4 Place Code, B7 Class மற்றும் R1:R24 வெற்றிபெற்ற ஜாக்கி பெயர்களை எடிட் செய்த பின் இங்கே பதிவேற்றவும். Col L-ல் AM ஸ்கோர் மற்றும் Col M-Q-ல் காரணங்கள் உருவாக்கப்படும்.")

    uploaded_file = st.file_uploader(
        "📂 திருத்தப்பட்ட Multi-Sheet Excel (.xlsx) ஃபைலை இங்கே பதிவேற்றவும்:",
        type=["xlsx"],
        key="tab2_file_uploader"
    )

    if uploaded_file is not None:
        if st.button("⚡ Process AM Scoring & Highlighting", type="primary", key="btn_step2"):
            with st.spinner("🔄 AM ஸ்கோர்கள் கணக்கிடப்பட்டு, அனைத்து ஷீட்டுகளும் புதுப்பிக்கப்படுகின்றன..."):
                try:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                        tmp.write(uploaded_file.read())
                        temp_in = tmp.name

                    wb = openpyxl.load_workbook(temp_in, data_only=False)
                    process_step2_edited_excel(wb)
                    wb.save(temp_in)

                    with open(temp_in, "rb") as f:
                        final_data = f.read()

                    try: os.remove(temp_in)
                    except Exception: pass

                    base_name = os.path.splitext(uploaded_file.name)[0]
                    final_filename = f"{base_name}_AM_FINAL.xlsx"

                    st.success("🎉 இறுதி AM PRO ஸ்கோர் அறிக்கை வெற்றிகரமாகத் தயாராகிவிட்டது!")
                    st.download_button(
                        label=f"📥 Download Final Excel ({final_filename})",
                        data=final_data,
                        file_name=final_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"Error: {str(e)}")